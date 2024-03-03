from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook as lw
from PIL import ImageTk, Image


def drop():
    menu.tk_popup(1050, 520)


def click(option):
    select_label1.config(text=option)
    login_button1.config(state=NORMAL)
    login_button1.focus_set()


def Back2(*event):
    t1.set('--Enter Username--'.center(46, ' '))
    t2.set('--Enter Password--'.center(46, ' '))
    login_button1.focus_set()
    password_entry.config(show='')
    notebook.select(0)


def username(*event):
    if username_entry.get() == '--Enter Username--'.center(46, ' '):
        username_entry.delete(0, END)
    username_entry.focus()


def password(*event):
    if password_entry.get() == '--Enter Password--'.center(46, ' '):
        password_entry.delete(0, END)
    password_entry.config(show='*')
    password_entry.focus()


def enter(*event):
    if select_label1.cget('text') == 'Student':
        student_page()
    elif select_label1.cget('text') == 'Faculty':
        faculty_page()
    elif select_label1.cget('text') == 'Admin':
        admin_page()


def check_pay_status(i, x):
    if x.cell(i, 5).value == 163000:
        s = 'Not Paid'
    elif x.cell(i, 5).value == 81500:
        s = '50% Paid'
    elif x.cell(i, 5).value == 0:
        s = 'Fully Paid'
    else:
        s = 'Others'
    return s


def login1(*event):
    if select_label1.cget('text') == 'Anonymous':
        notebook.select(2)
    else:
        if select_label1.cget('text') == 'Admin':
            username_entry.delete(0, END)
            username_entry.insert(0, 'Admin')
        notebook.select(1)
        return

    def Back3(*event):
        notebook.select(0)
        label1.config(text='--- Select Branch ---')
        label2.config(text='    --- Select Semester ---')
        label3.grid_forget()

    def disp_fee_details():
        if label1.cget('text') != '--- Select Branch ---' and label2.cget('text') != '    --- Select Semester ---':
            label3.grid(row=4, column=5, rowspan=5, columnspan=40)

    def click_branch(*event):
        label1.config(text='{:<20s}'.format(clicked_branch.get()))
        clicked_branch.set('')
        disp_fee_details()

    def click_sem(*event):
        label2.config(text=5 * ' ' + clicked_sem.get())
        clicked_sem.set('')
        disp_fee_details()
    bg_anonymous = Label(mainframe3, image=bg_image)
    bg_anonymous.grid(row=0, column=0, rowspan=10, columnspan=65)

    button_back3 = Button(mainframe3, image=back2_image, command=Back3, bd=0, bg='black', activebackground='black')
    button_back3.grid(row=0, column=2)

    label1 = Label(mainframe3, text='--- Select Branch ---', font=('Calibri', 20, 'bold'), width=25, bd=0, bg='#FFFFFF')
    label1.grid(row=3, column=5, pady=47)

    branches = ['CSE', 'CIVIL', 'CHEMICAL', 'BIOTECH', 'EEE', 'IT', 'CIC', 'AIDS', 'CSM', 'MECHANICAL', 'ECE']

    clicked_branch = StringVar()
    clicked_branch.set('')
    drop_branch = OptionMenu(mainframe3, clicked_branch, *branches, command=click_branch)
    drop_branch.config(bg='#FFFFFF', borderwidth=0)
    drop_branch.grid(row=3, column=6)

    label2 = Label(mainframe3, text='    --- Select Semester ---', font=('Calibri', 20, 'bold'), width=25, bd=0,
                   bg='#FFFFFF')
    label2.grid(row=3, column=36)

    semesters = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII']

    clicked_sem = StringVar()
    clicked_sem.set('')
    drop_sem = OptionMenu(mainframe3, clicked_sem, *semesters, command=click_sem)
    drop_sem.config(bg='#FFFFFF', borderwidth=0)
    drop_sem.grid(row=3, column=37)

    label3 = Label(mainframe3, image=fee_details, bg='yellow')


def student_page():
    def Back4(*event):
        B_sheet.cell(i, 6).value = 'No Notifications'
        xlsx_Database.save('Book 1st year.xlsx')
        notebook.select(1)
        password_entry.delete(0, END)

    def amount():
        s = 0
        var_e.set(0)
        var_f.set(0)

        if var_a.get() == 1:
            s += Tuttion_fee
        if var_b.get() == 1:
            s += Caution_deposit
        if var_c.get() == 1:
            s += Library_fee
        if var_d.get() == 1:
            s += Placement_training
        if s == 0:
            s = ''
        amount_entry.config(text=str(s))

    def amount2(var_s):
        var_a.set(0)
        var_b.set(0)
        var_c.set(0)
        var_d.set(0)
        var_s.set(0)
        amount_entry.config(text='')
        if var_e.get() == 1:
            amount_entry.config(text=str(Partial_fee))
        if var_f.get() == 1:
            amount_entry.config(text=str(Total_fee))

    def pay():
        global glob_cell
        if var.get() == 0:
            messagebox.showinfo('No Payment method selected', 'Please select a payment method')
            return
        if amount_entry.cget('text') == '':
            return
        fee = cell.value
        fee -= int(amount_entry.cget('text'))
        cell.value = fee
        xlsx_Database.save('Book 1st year.xlsx')
        if B_sheet.cell(i, 5).value == fee:
            messagebox.showinfo('Transaction Update', 'Fee Successfully Updated...')
        buttons_activate()

    B_sheet = xlsx_Database['Branches']
    rollno = username_entry.get()
    if rollno[:6] != '160121':
        messagebox.showerror('ERROR', 'Invalid Username')
        return
    for i in range(1, B_sheet.max_column+1):
        if rollno[6:9] == str(B_sheet.cell(2, i).value):
            Branch = B_sheet.cell(1, i).value
            break
    else:
        messagebox.showerror('ERROR', 'Invalid Username')
        return
    B_sheet = xlsx_Database[Branch]
    for i in range(4, B_sheet.max_row + 1):
        if rollno[9:] == str(B_sheet.cell(i, 2).value):
            if password_entry.get() == str(B_sheet.cell(i, 4).value):
                notebook.select(3)
                break
            else:
                messagebox.showerror('ERROR', 'Invalid Password')
                return
    else:
        messagebox.showerror('ERROR', 'Invalid Username')
        return
    cell = B_sheet.cell(i, 5)
    bg_student = Label(mainframe4, image=student_bg_image)
    bg_student.grid(row=0, column=0, rowspan=16, columnspan=10)

    notif_label = Label(mainframe4, text=B_sheet.cell(i, 6).value, font=('Inter', 20), bd=0, bg='#D9D9D9', fg='blue')
    notif_label.place(x=350, y=168)

    feesummary_name = Label(mainframe4, bg='#FFFFFF', font=('Inter', 16, 'bold'), text='', bd=0)
    feesummary_text = Text(mainframe4, bg='#FFFFFF', font=('Inter', 16, 'bold'), height=8, width=30, bd=0)

    button_back4 = Button(mainframe4, image=back2_image, command=Back4, bd=0, bg='black', activebackground='black')
    button_back4.grid(row=0, column=0, pady=3)

    payment_frame = Frame(mainframe4, bg='#FFDC26')

    var_a = IntVar()
    var_b = IntVar()
    var_c = IntVar()
    var_d = IntVar()
    var_e = IntVar()
    var_f = IntVar()

    TF_button = Checkbutton(mainframe4, variable=var_a, onvalue=1, offvalue=0, text='Tuition Fee', font=('Inder', 17), bg='#FFDC26', activebackground='#FFDC26', command=amount)
    TF_button.place(x=950, y=330)

    CD_button = Checkbutton(mainframe4, variable=var_b, onvalue=1, offvalue=0, text='Caution Deposit', font=('Inder', 17), bg='#FFDC26', activebackground='#FFDC26', command=amount)
    CD_button.place(x=950, y=360)

    LF_button = Checkbutton(mainframe4, variable=var_c, onvalue=1, offvalue=0, text='Library Fee', font=('Inder', 17), bg='#FFDC26', activebackground='#FFDC26', command=amount)
    LF_button.place(x=950, y=390)

    PT_button = Checkbutton(mainframe4, variable=var_d, onvalue=1, offvalue=0, text='Placement Training', font=('Inder', 17), bg='#FFDC26', activebackground='#FFDC26', command=amount)
    PT_button.place(x=950, y=420)

    Partial_fee_button = Checkbutton(mainframe4, variable=var_e, onvalue=1, offvalue=0, text='50% Fee', font=('Inder', 17), bg='#FFDC26', activebackground='#FFDC26', command= lambda: amount2(var_f))
    Partial_fee_button.place(x=950, y=450)

    Total_fee_button = Checkbutton(mainframe4, variable=var_f, onvalue=1, offvalue=0, text='Total Fee', font=('Inder', 17), bg='#FFDC26', activebackground='#FFDC26', command=lambda: amount2(var_e))
    Total_fee_button.place(x=950, y=480)

    amount_label2 = Label(payment_frame, text='Amount', font=('Inter', 17, 'bold'), bd=0, bg='#FFDC26')
    amount_label2.grid(row=1, column=1, pady=8)

    image_label2 = Label(payment_frame, image=payment_entry_image2, bd=0, bg='#FFDC26')
    image_label2.grid(row=2, column=1, pady=8)

    amount_entry = Label(payment_frame, width=10, font=('Inter', 17, 'bold'), bg='#FFFFFF', bd=0)
    amount_entry.grid(row=2, column=1, pady=8)

    button_pay = Button(payment_frame, text='Pay', font=('Inter', 12, 'bold'), bg='#37EB5E', fg='#FF0000', activebackground='#37EB5E', command=pay)
    button_pay.grid(row=2, column=2, pady=8)

    button_req = Button(payment_frame, text='Request for concession', font=('Inter', 16, 'bold'), bg='#ABFC00', activebackground='#ABFC00')
    button_req.grid(row=3, column=1, columnspan=2, pady=0)

    var = IntVar()

    radiobutton1 = Radiobutton(mainframe4, variable=var, value=1, text='NET BANKING', font=('Inder', 24), bg='#FFFFFF', activebackground='#FFFFFF')
    radiobutton2 = Radiobutton(mainframe4, variable=var, value=2, text='CREDIT/DEBIT CARD', font=('Inder', 24), bg='#FFFFFF', activebackground='#FFFFFF')
    radiobutton3 = Radiobutton(mainframe4, variable=var, value=3, image=UPI_image, bg='#FFFFFF', activebackground='#FFFFFF')

    def buttons_activate():
        var_a.set(0)
        var_b.set(0)
        var_c.set(0)
        var_d.set(0)
        var_e.set(0)
        var_f.set(0)

        amount_entry.config(text='')
        j = Total_fee - B_sheet.cell(i, 5).value
        if j == Total_fee:
            TF_button.config(state=DISABLED); CD_button.config(state=DISABLED); LF_button.config(state=DISABLED); PT_button.config(state=DISABLED); Partial_fee_button.config(state=DISABLED); Total_fee_button.config(state=DISABLED)
        elif j == Partial_fee:
            TF_button.config(state=DISABLED)
            CD_button.config(state=DISABLED)
            LF_button.config(state=DISABLED)
            PT_button.config(state=DISABLED)
            Total_fee_button.config(state=DISABLED)
        elif j != 0:
            if j == Tuttion_fee:
                TF_button.config(state=DISABLED)
            elif j == Caution_deposit:
                CD_button.config(state=DISABLED)
            elif j == Library_fee:
                LF_button.config(state=DISABLED)
            elif j == Placement_training:
                PT_button.config(state=DISABLED)
            elif j == Tuttion_fee + Caution_deposit:
                TF_button.config(state=DISABLED)
                CD_button.config(state=DISABLED)
            elif j == Tuttion_fee + Library_fee:
                TF_button.config(state=DISABLED)
                LF_button.config(state=DISABLED)
            elif j == Tuttion_fee + Placement_training:
                TF_button.config(state=DISABLED)
                PT_button.config(state=DISABLED)
            elif j == Caution_deposit + Library_fee:
                CD_button.config(state=DISABLED)
                LF_button.config(state=DISABLED)
            elif j == Caution_deposit + Placement_training:
                CD_button.config(state=DISABLED)
                PT_button.config(state=DISABLED)
            elif j == Library_fee + Placement_training:
                LF_button.config(state=DISABLED)
                PT_button.config(state=DISABLED)
            elif j == Tuttion_fee + Caution_deposit + Library_fee:
                TF_button.config(state=DISABLED)
                CD_button.config(state=DISABLED)
                LF_button.config(state=DISABLED)
            elif j == Tuttion_fee + Caution_deposit + Placement_training:
                TF_button.config(state=DISABLED)
                CD_button.config(state=DISABLED)
                PT_button.config(state=DISABLED)
            elif j == Tuttion_fee + Library_fee + Placement_training:
                TF_button.config(state=DISABLED)
                LF_button.config(state=DISABLED)
                PT_button.config(state=DISABLED)
            elif j == Caution_deposit + Library_fee + Placement_training:
                CD_button.config(state=DISABLED)
                LF_button.config(state=DISABLED)
                PT_button.config(state=DISABLED)
            Partial_fee_button.config(state=DISABLED)
            Total_fee_button.config(state=DISABLED)
        feesummary_text.config(state=NORMAL)
        feesummary_text.insert('1.0', '\n{:<10s} :  {}\n\n{:<9s} :  {}\n{:^6}\n\n{:<9s} :  {}\n{:^6s}'.format('Total Fee', str(Total_fee), 'Amount', str(j), 'Paid', 'Amount', str(B_sheet.cell(i, 5).value), ' due'))
        feesummary_text.config(state=DISABLED)
        feesummary_text.grid(row=7, column=1, sticky=S + W)

    feesummary_name.config(text='{:<11s} :  {}'.format('Name', B_sheet.cell(i, 3).value))
    feesummary_name.grid(row=7, column=1, sticky=N+W, pady=17, columnspan=5)
    buttons_activate()
    radiobutton1.place(x=130, y=698)
    radiobutton2.place(x=600, y=698)
    radiobutton3.place(x=1150, y=668)

    payment_frame.grid(row=6, column=8, rowspan=2, sticky=S)


def faculty_page():
    def Back5():
        xlsx_Database.save('Book 1st year.xlsx')
        notebook.select(1)
        password_entry.delete(0, END)

    def multiple_yview(*args):
        Student_Rollnos.yview(*args)
        Pay_status.yview(*args)

    def category(*event):
        if clicked_status.get() == 'Custom'.center(26, ' '):
            search_entry.place(x=120, y=85)
            search_entry.focus()
            search_button.config(state=NORMAL)
            return
        else:
            search_entry.delete(0, END)
            search_entry.place_forget()
            search_button.config(state=DISABLED)
        Student_Rollnos.config(state=NORMAL)
        Pay_status.config(state=NORMAL)
        Student_Rollnos.delete('1.0', 'end-1c')
        Pay_status.delete('1.0', 'end-1c')
        for i in range(4, Branch.max_row + 1):
            s = check_pay_status(i, Branch)
            if s in clicked_status.get() or 'All' in clicked_status.get() or 'Select Pay Status' in clicked_status.get():
                Student_Rollnos.insert(END, '160121' + code + str(Branch.cell(i, 2).value) + '\n\n')
                if s == 'Others':
                    s = str(Branch.cell(i, 5).value)
                    s = s[:-3] + ',' + s[-3:]
                    if len(s) == 7:
                        s = s[0] + ',' + s[1:]
                Pay_status.insert(END, s + '\n\n')
        Student_Rollnos.config(state=DISABLED)
        Pay_status.config(state=DISABLED)

    def search():
        if search_entry.get()[:6] == '160121' and search_entry.get()[6:9] == code:
            Student_Rollnos.config(state=NORMAL)
            Pay_status.config(state=NORMAL)
            for i in range(4, Branch.max_row + 1):
                if search_entry.get()[9:] == Branch.cell(i, 2).value:
                    Student_Rollnos.delete('1.0', 'end-1c')
                    Pay_status.delete('1.0', 'end-1c')
                    s = check_pay_status(i, Branch)
                    if s == 'Others':
                        s = str(Branch.cell(i, 5).value)
                        s = s[:-3] + ',' + s[-3:]
                        if len(s) == 7:
                            s = s[0] + ',' + s[1:]
                    Student_Rollnos.insert(END, '160121' + code + str(Branch.cell(i, 2).value) + '\n\n')  # 733
                    Pay_status.insert(END, s + '\n\n')
                    break
            else:
                messagebox.showerror('ERROR', 'Roll number not found')
            Student_Rollnos.config(state=DISABLED)
            Pay_status.config(state=DISABLED)
        else:
            messagebox.showerror('ERROR', 'Roll number not found')

    def custom(*event):
        clicked_student.set(pay_Status2[4])

    def message_box_activate(*event):
        if message_box.get('1.0', 'end-1c') == 'Enter message here...'.ljust(50, ' '):
            message_box.delete('1.0', END)
        message_box.focus()

    def selected2(*event):
        if clicked_student.get() == pay_Status2[4]:
            to_entry.focus()
        else:
            to_entry.delete(0, END)
            message_box_activate()

    def send_mssg():
        if message_box.get("1.0", "end-1c") == 'Enter message here...'.ljust(50, ' ') or message_box.get("1.0", "end-1c") == '':
            messagebox.showinfo('No Message', 'Enter a message')
            return
        if clicked_student.get() == pay_Status2[4]:
            if to_entry.get()[:6] == '160121' and to_entry.get()[6:9] == code:
                for i in range(4, Branch.max_row + 1):
                    if to_entry.get()[9:] == Branch.cell(i, 2).value:
                        Branch.cell(i, 6).value = message_box.get('1.0', 'end-1c')
                        xlsx_Database.save('Book 1st year.xlsx')
                        messagebox.showinfo('Information', 'Message sent')
                        break
                else:
                    messagebox.showerror('ERROR', 'Roll number not found')
            else:
                messagebox.showerror('ERROR', 'Roll number not found')
        else:
            for i in range(4, Branch.max_row + 1):
                s = check_pay_status(i, Branch)
                if s in clicked_student.get():
                    Branch.cell(i, 6).value = message_box.get('1.0', 'end-1c')
            xlsx_Database.save('Book 1st year.xlsx')
            messagebox.showinfo('Information', 'Message sent')

    def delete_mssg():
        if message_box.get('1.0', 'end-1c') != 'Enter message here...'.ljust(50, ' '):
            message_box.delete('1.0', END)

    Faculty = xlsx_Database['Faculty']
    username = username_entry.get()
    for i in range(2, Faculty.max_row + 1):
        if username == str(Faculty.cell(i, 1).value):
            if password_entry.get() == str(Faculty.cell(i, 2).value):
                notebook.select(4)
                Branch = Faculty.cell(i, 3).value
                code = str(Faculty.cell(i, 4).value)
                break
            else:
                messagebox.showerror('ERROR', 'Invalid Password')
                return
    else:
        messagebox.showerror('ERROR', 'Invalid Username')
        return
    Branch = xlsx_Database[Branch]
    bg_faculty = Label(mainframe5, image=faculty_bg_image)
    bg_faculty.grid(row=0, column=0, rowspan=40, columnspan=20)

    button_back4 = Button(mainframe5, image=back2_image, command=Back5, bd=0, bg='black', activebackground='black')
    button_back4.place(x=30, y=10)

    search_entry = Entry(mainframe5, width=20, bd=5, bg='Gray', font=('Inter', 18), justify=CENTER, relief=RIDGE)

    search_button = Button(mainframe5, image=search_bg, bd=0, command=search, state=DISABLED)
    search_button.place(x=390, y=85)

    pay_Status = ['All'.center(30, ' '), 'Fully Paid'.center(27, ' '), 'Not Paid'.center(27, ' '), '50% Paid'.center(24, ' '), 'Others'.center(27, ' '), 'Custom'.center(26, ' ')]

    clicked_status = StringVar()
    clicked_status.set('Select Pay Status'.center(20, ' '))
    drop_status = OptionMenu(mainframe5, clicked_status, *pay_Status, command=category)
    drop_status.config(bg='#FFFFFF', borderwidth=0, font=('Inder', 10), width=15)
    drop_status.place(x=650, y=92)

    text_scroll = Scrollbar(mainframe5)
    text_scroll.grid(row=6, column=8, rowspan=29, sticky=N + S + W, pady=10)

    Student_Rollnos = Text(mainframe5, width=12, height=18, pady=10, font=('Inter', 20), yscrollcommand=text_scroll.set, wrap='none', bd=0)
    Student_Rollnos.place(x=173, y=130)

    Pay_status = Text(mainframe5, width=9, height=18, pady=10, font=('Inter', 20), yscrollcommand=text_scroll.set, wrap='none', bd=0)
    Pay_status.place(x=450, y=130)

    text_scroll.config(command=multiple_yview)

    from_label = Label(mainframe5, text='From', font=('Inter', 24), bd=0, bg='#FFDC26')
    from_label.place(x=900, y=180)

    to_label = Label(mainframe5, text='To', font=('Inter', 24), bd=0, bg='#FFDC26')
    to_label.place(x=900, y=240)

    from_image = Label(mainframe5, image=from_or_to_entry_image, bd=0)
    from_image.place(x=1000, y=180)

    to_image = Label(mainframe5, image=from_or_to_entry_image, bd=0)
    to_image.place(x=1000, y=240)

    from_entry = Label(mainframe5, width=23, bd=0, font=('Inter', 18), justify=CENTER, text=username, bg='#FFFFFF')
    from_entry.place(x=1035, y=189)

    to_entry = Entry(mainframe5, width=23, bd=0, font=('Inter', 18), justify=CENTER)
    to_entry.place(x=1040, y=249)
    to_entry.bind('<Button-1>', custom)
    to_entry.bind('<Tab>', message_box_activate)
    to_entry.bind('<Return>', message_box_activate)

    message_box = Text(mainframe5, width=42, height=3, font=('Inter', 18))
    message_box.place(x=890, y=330)
    message_box.insert(END, 'Enter message here...'.ljust(50, ' '))
    message_box.bind('<Button-1>', message_box_activate)

    pay_Status2 = ['Fully Paid'.center(103, ' '), 'Not Paid'.center(103, ' '), '50% Paid'.center(100, ' '), 'Others'.center(103, ' '), 'Custom'.center(101, ' ')]

    clicked_student = StringVar()
    clicked_student.set('-- Select Student --'.center(35, ' '))
    select_student = OptionMenu(mainframe5, clicked_student, *pay_Status2, command=selected2)
    select_student.config(bg='#FFFFFF', borderwidth=5, relief=RIDGE, font=('Inder', 18), width=25, activebackground='#FFFFFF', justify=RIGHT)
    select_student.place(x=915, y=473)

    button_send = Button(mainframe5, text='SEND', bg='#43EB68', width=7, font=('Inder', 16, 'bold'), command=send_mssg)
    button_send.place(x=1320, y=480)

    button_delete = Button(mainframe5, text='DELETE', bg='#FF0000', width=7, font=('Inder', 16, 'bold'),
                           command=delete_mssg)
    button_delete.place(x=1320, y=550)

    category()


def admin_page():
    Admin = xlsx_Database['Admin']
    username = username_entry.get()
    if username == str(Admin.cell(2, 1).value):
            if password_entry.get() == str(Admin.cell(2, 2).value):
                notebook.select(5)
            else:
                messagebox.showerror('ERROR', 'Invalid Password')
                return
    else:
        messagebox.showerror('ERROR', 'Invalid Username')
        return

    def Back6():
        xlsx_Database.save('Book 1st year.xlsx')
        notebook.select(1)
        password_entry.delete(0, END)

    def continue1(*event):
        if name_entry.get() == '':
            messagebox.showwarning('WARNING', 'Please enter the Name')
            return
        Branches = xlsx_Database['Branches']
        if branch_entry.get() == '':
            messagebox.showwarning('WARNING', 'Please enter the Branch')
            return
        for i in range(1, Branches.max_column + 1):
            if branch_entry.get().upper() == Branches.cell(1, i).value:
                rollno_entry.config(state=NORMAL)
                Branch = branch_entry.get().upper()
                code = str(Branches.cell(2, i).value)
                rollno_entry.insert(0, '160121' + code + str(
                    int(xlsx_Database[Branch].cell(xlsx_Database[Branch].max_row, 2).value) + 1))
                if len(rollno_entry.get()) == 11:
                    rollno_entry.insert(9, '0')
                rollno_entry.config(state=DISABLED)
                break
        else:
            messagebox.showwarning('WARNING', 'Please input valid branch')
            return
        button_continue.place_forget()
        name_entry.config(state=DISABLED)
        branch_entry.config(state=DISABLED)
        rollno_label.place(x=200, y=520)
        rollno_entry.place(x=350, y=520)
        paystatus_label.place(x=200, y=620)
        paystatus_entry.place(x=350, y=625)
        button_back.place(x=460, y=700)
        button_add.place(x=545, y=700)

    def add(Branch):
        Branch = xlsx_Database[Branch]
        row = Branch.max_row
        Branch.cell(row + 1, 1).value = Branch.cell(row, 1).value + 1
        Branch.cell(row + 1, 2).value = str(rollno_entry.get()[9:12])
        Branch.cell(row + 1, 3).value = str(name_entry.get().upper())
        Branch.cell(row + 1, 4).value = str(rollno_entry.get())
        Branch.cell(row + 1, 5).value = int(paystatus_entry.get())
        Branch.cell(row + 1, 6).value = 'No Notifications'
        xlsx_Database.save('Book 1st year.xlsx')
        messagebox.showinfo('Information', ' Student record successfully added')
        cancel()

    def cancel():
        button_continue.place(x=480, y=510)
        name_entry.config(state=NORMAL)
        branch_entry.config(state=NORMAL)
        name_entry.delete(0, END)
        branch_entry.delete(0, END)
        rollno_entry.config(state=NORMAL)
        rollno_entry.delete(0, END)
        rollno_entry.config(state=DISABLED)
        rollno_label.place_forget()
        rollno_entry.place_forget()
        paystatus_label.place_forget()
        paystatus_entry.place_forget()
        button_back.place_forget()
        button_add.place_forget()
        name_label.focus_set()

    def roll_box(*event):
        if enter_roll.get() == '--Enter Roll number--'.center(50, ' '):
            t1.set('')

    def check(*event):
        if enter_roll.get() == '--Enter Roll number--'.center(50, ' ') or enter_roll.get() == '':
            messagebox.showwarning('Warning', 'Please enter a Roll number')
            return
        if enter_roll.get()[:6] == '160121' and len(enter_roll.get()) == 12:
            code = enter_roll.get()[6:9]
            Branches = xlsx_Database['Branches']
            for i in range(1, Branches.max_column + 1):
                if code == str(Branches.cell(2, i).value):
                    branch = Branches.cell(1, i).value
                    Branch = xlsx_Database[branch]
                    for j in range(4, Branch.max_row + 1):
                        if enter_roll.get()[9:12] == str(Branch.cell(j, 2).value):
                            text_frame.config(state=NORMAL)
                            text_frame.delete('1.0', 'end-1c')
                            text_frame.insert(END, 'Name  : {}\nBranch: {}\nPay Status: '.format(
                                str(Branch.cell(j, 3).value), branch))
                            if Branch.cell(j, 5).value == Total_fee:
                                text_frame.insert(END, '\nNot Paid')
                            else:
                                other = {Total_fee / 2: ['50% fee'], 0: ['Full fee'], Tuttion_fee: ['Tuition fee'],
                                         Caution_deposit: ['Caution deposit'], Library_fee: ['Library fee'],
                                         Placement_training: ['Placement training'],
                                         Tuttion_fee + Caution_deposit: ['Tuition fee', 'Caution deposit'],
                                         Tuttion_fee + Library_fee: ['Tuition fee', 'Library fee'],
                                         Tuttion_fee + Placement_training: ['Tuition fee', 'Placement training'],
                                         Caution_deposit + Library_fee: ['Caution deposit', 'Library fee'],
                                         Caution_deposit + Placement_training: ['Caution deposit',
                                                                                'Placement training'],
                                         Tuttion_fee + Caution_deposit + Library_fee: ['Tuition fee', 'Caution deposit',
                                                                                       'Library fee'],
                                         Tuttion_fee + Caution_deposit + Placement_training: ['Tuition fee',
                                                                                              'Caution deposit',
                                                                                              'Placement training'],
                                         Tuttion_fee + Library_fee + Placement_training: ['Tuition fee', 'Library fee',
                                                                                          'Placement training'],
                                         Caution_deposit + Library_fee + Placement_training: ['Caution deposit',
                                                                                              'Library fee',
                                                                                              'Placement training']}
                                pay = other[Branch.cell(j, 5).value]
                                for k in pay:
                                    text_frame.insert(END, '\nPaid ' + k)
                            text_frame.config(state=DISABLED)
                            break
                    break
            else:
                messagebox.showerror('ERROR', 'Roll number not found')
        else:
            messagebox.showerror('ERROR', 'Roll number not found')

    bg_faculty = Label(mainframe6, image=admin_bg_image)
    bg_faculty.grid(row=0, column=0, rowspan=10, columnspan=10)

    button_back5 = Button(mainframe6, image=back2_image, command=Back6, bd=0, bg='black', activebackground='black')
    button_back5.place(x=30, y=15)

    name_label = Label(mainframe6, text='Name', font=('Inter', 20), bd=0, bg='#FFFFFF')
    name_label.place(x=200, y=320)

    name_entry = Entry(mainframe6, font=('Inter', 20), bd=0, bg='#D9D9D9', width=20)
    name_entry.place(x=350, y=320)

    branch_label = Label(mainframe6, text='Branch', font=('Inter', 20), bd=0, bg='#FFFFFF')
    branch_label.place(x=200, y=420)

    branch_entry = Entry(mainframe6, font=('Inter', 20), bd=0, bg='#D9D9D9', width=20)
    branch_entry.place(x=350, y=420)
    branch_entry.bind('<Return>', continue1)

    button_continue = Button(mainframe6, text='Continue', font=('Inter', 16, 'bold'), bg='#37EB5E', width=10, command=continue1)
    button_continue.place(x=480, y=510)

    rollno_label = Label(mainframe6, text='Roll no.', font=('Inter', 20), bd=0, bg='#FFFFFF')

    rollno_entry = Entry(mainframe6, font=('Inter', 20), bd=0, bg='#D9D9D9', width=20, state=DISABLED)

    paystatus_label = Label(mainframe6, text='Pay   \nstatus', font=('Inter', 20), bd=0, bg='#FFFFFF')

    paystatus_entry = Entry(mainframe6, font=('Inter', 20), bd=0, bg='#D9D9D9', width=20)
    paystatus_entry.insert(0, Total_fee)
    paystatus_entry.config(state=DISABLED)

    button_add = Button(mainframe6, text='Add', font=('Inter', 16, 'bold'), bg='#37EB5E', width=5, command=lambda: add(branch_entry.get().upper()))

    button_back = Button(mainframe6, text='Back', font=('Inter', 16, 'bold'), bg='#FF0000', width=5, command=cancel)

    t1 = StringVar()
    t1.set('--Enter Roll number--'.center(50, ' '))
    enter_roll = Entry(mainframe6, textvariable=t1, font=('Inter', 20), bg='#FFFFFF', width=30, justify=CENTER)
    enter_roll.place(x=887, y=320)
    enter_roll.bind('<Button-1>', roll_box)
    enter_roll.bind('<Return>', check)

    check_pay_button = Button(mainframe6, text='Pay status', font=('Inter', 16), bg='#FFFFFF', width=37, bd=5,
                              command=check)
    check_pay_button.place(x=887, y=420)

    text_frame = Text(mainframe6, font=('Inter', 18), bg='#FFFFFF', width=38, height=6, state=DISABLED, bd=7,
                      relief=RIDGE)
    text_frame.place(x=860, y=580)


def close_root():
    xlsx_Database.save('Book 1st year.xlsx')
    root.destroy()

root = Tk()
root.title('CBIT Fee Management')
root.state('zoomed')
root.config(bg='#FFFFFF')
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

notebook = ttk.Notebook(root)
notebook.pack()

style = ttk.Style()
style.layout('TNotebook.Tab', [])

cover_image = ImageTk.PhotoImage(Image.open('cbit cover 2.jpg').resize((screen_width,screen_height), Image.Resampling.LANCZOS))
logo_image1 = ImageTk.PhotoImage(Image.open('cbit logo.png').resize((160, 200), Image.Resampling.LANCZOS))
logo_image2 = ImageTk.PhotoImage(Image.open('cbit logo.png').resize((320, 400), Image.Resampling.LANCZOS))
entry_image = ImageTk.PhotoImage(Image.open('trial box.png').resize((515, 60), Image.Resampling.LANCZOS))
bottom_content = ImageTk.PhotoImage(Image.open('bottom content.png').resize((screen_width//6,screen_height//6),Image.Resampling.LANCZOS))
drop_down = ImageTk.PhotoImage(Image.open('drop button.jpeg'))
back_image = ImageTk.PhotoImage(Image.open('back.png'))
back2_image = ImageTk.PhotoImage(Image.open('back 2.png'))
bg_image = ImageTk.PhotoImage(Image.open('bg pic.jpeg').resize((screen_width,screen_height), Image.Resampling.LANCZOS))
fee_details = ImageTk.PhotoImage(Image.open('Fee pic.png'))
student_bg_image = ImageTk.PhotoImage(Image.open('student_bg pic.jpeg').resize((screen_width,screen_height),Image.Resampling.LANCZOS))
payment_entry_image1 = ImageTk.PhotoImage(Image.open('yellowbg_entry_box.png').resize((310, 40),Image.Resampling.LANCZOS))
payment_entry_image2 = ImageTk.PhotoImage(Image.open('yellowbg_entry_box.png').resize((200, 40),Image.Resampling.LANCZOS))
UPI_image = ImageTk.PhotoImage(Image.open('UPI image.png'))#.resize((200, 40),Image.Resampling.LANCZOS))
faculty_bg_image = ImageTk.PhotoImage(Image.open('faculty_bg pic.jpeg').resize((screen_width, screen_height-50),Image.Resampling.LANCZOS))
search_bg = ImageTk.PhotoImage(Image.open('search_button.jpeg'))
from_or_to_entry_image = ImageTk.PhotoImage(Image.open('yellowbg_entry_box.png').resize((400, 45), Image.Resampling.LANCZOS))
admin_bg_image = ImageTk.PhotoImage(Image.open('admin_bg pic.jpeg').resize((screen_width, screen_height), Image.Resampling.LANCZOS))

xlsx_Database = lw('Book 1st year.xlsx')

Total_fee, Partial_fee, Tuttion_fee, Caution_deposit, Library_fee, Placement_training = 163000, 81500, 134000, 7500, 1250, 20250


mainframe1 = Frame(notebook, bg='#FFFFFF')
mainframe1.pack()

top_frame1 = Frame(mainframe1)
top_frame1.pack(side=TOP)

middle_frame1 = Frame(mainframe1, bg='#FFFFFF')
middle_frame1.pack(anchor='w', padx=150, pady=50, expand=1)

bottom_frame1 = Frame(mainframe1, bg='light sky blue')
bottom_frame1.pack(side=BOTTOM)

cover_label1 = Label(top_frame1, image=cover_image, height=screen_height//2.5)
cover_label1.pack()

logo_label1 = Label(middle_frame1, image=logo_image1, bg='#FFFFFF')
logo_label1.grid(row=0, column=0, rowspan=4, padx=100)

clicked = StringVar()
clicked.set('')

image_label1 = Label(middle_frame1, image=entry_image, bd=0)
image_label1.grid(row=1, column=1)

select_label1 = Label(middle_frame1, text='Select User Type', font=('Calibri',20,'bold'), width=33, bd=0, bg='#FFFFFF')
select_label1.grid(row=1, column=1)

drop1 = Button(middle_frame1, bg='#FFFFFF', borderwidth=0, command=drop, image=drop_down)
drop1.grid(row=1, column=2, padx=10)

menu = Menu(root, tearoff=0, bg='#FFFFFF')
# options = ['Anonymous', 'Student', 'Faculty', 'Admin']  # , 'HoD', 'Principal']
menu.add_command(label='Anonymous', command=lambda: click('Anonymous'))
menu.add_command(label='Student', command=lambda: click('Student'))
menu.add_command(label='Faculty', command=lambda: click('Faculty'))
menu.add_command(label='Admin', command=lambda: click('Admin'))

login_button1 = Button(middle_frame1, text='LOGIN', font=('Calibri',15,'bold'), bd=0, bg='#FFFFFF', activebackground='#FFFFFF', state=DISABLED, image=entry_image, command=login1)
login_button1.grid(row=2, column=1, sticky=W)

button_help = Button(bottom_frame1, text='Help', font=('Calibri',15,'bold'), bd=0, bg='light sky blue', activebackground='light sky blue')
button_help.grid(row=0, column=0, sticky=W+N, padx=50, pady=20)

button_explore = Button(bottom_frame1, text='Explore', font=('Calibri',15,'bold'), bd=0, bg='light sky blue', activebackground='light sky blue')
button_explore.grid(row=0, column=1, sticky=W+N, padx=200, pady=20)

button_contactus = Button(bottom_frame1, text='Contact Us', font=('Calibri',15,'bold'), bd=0, bg='light sky blue', activebackground='light sky blue')
button_contactus.grid(row=0, column=2, sticky=W+N, padx=100, pady=20)

bottom_text1 = Label(bottom_frame1, image=bottom_content, bd=0)
bottom_text1.grid(row=0, column=4, padx=300)

notebook.add(mainframe1)

mainframe2 = Frame(notebook, bg='#FFFFFF')
mainframe2.pack(expand=1)

top_frame2 = Frame(mainframe2)
top_frame2.pack(side=TOP)

middle_frame2 = Frame(mainframe2, bg='#FFFFFF')
middle_frame2.pack(anchor='w', padx=10, pady=10,expand=1)

cover_label2 = Label(top_frame2, image=cover_image, height=screen_height//2.5)
cover_label2.pack()

button_back2 = Button(middle_frame2, image=back_image, command=Back2, bd=0, bg='#FFFFFF', activebackground='#FFFFFF')
button_back2.grid(row=0, column=0, sticky=N)

logo_label2 = Label(middle_frame2, image=logo_image2, bg='#FFFFFF')
logo_label2.grid(row=0, column=1, rowspan=7, padx=500, pady=20)

image_label1 = Label(middle_frame2, image=entry_image, bd=0)
image_label1.grid(row=2, column=1)

image_label2 = Label(middle_frame2, image=entry_image, bd=0)
image_label2.grid(row=3, column=1)

t1 = StringVar()
t2 = StringVar()
t1.set('--Enter Username--'.center(46, ' '))
t2.set('--Enter Password--'.center(46, ' '))

username_entry = Entry(middle_frame2, textvariable=t1, font=('Calibri', 25, 'bold'), width=27, bd=0)
username_entry.grid(row=2, column=1)

password_entry = Entry(middle_frame2, textvariable=t2, font=('Calibri', 25, 'bold'), width=27, bd=0, show='')
password_entry.grid(row=3, column=1)

Clicked00 = login_button1.bind('<Return>', login1)
Clicked01 = button_back2.bind('<Tab>', username)
Clicked11 = username_entry.bind('<Button-1>', username)
Clicked12 = username_entry.bind('<Tab>', password)
Clicked13 = username_entry.bind('<Return>', password)
Clicked21 = password_entry.bind('<Button-1>', password)
# Clicked22 = password_entry.bind('<Tab>')
Clicked23 = password_entry.bind('<Return>', enter)

notebook.add(mainframe2)

mainframe3 = Frame(notebook)
mainframe3.pack()

notebook.add(mainframe3)

mainframe4 = Frame(notebook, bg='#FFFFFF')
mainframe4.pack()

notebook.add(mainframe4)

mainframe5 = Frame(notebook)
mainframe5.pack()

notebook.add(mainframe5)

mainframe6 = Frame(notebook, bg='#FFFFFF')
mainframe6.pack()

notebook.add(mainframe6)

root.protocol('WM_DELETE_WINDOW', close_root)
root.mainloop()
